import openpyxl

# Carregar a planilha
workbook = openpyxl.load_workbook(r'C:\Users\João Pedro Cordeiro\Downloads\Cópia-de-Funcionários-Nova_1_.xlsx')
sheet = workbook.active  # Ou use workbook['Nome_da_Sheet']

# Ler fórmulas
for row in sheet.iter_rows(min_row=1, max_col=sheet.max_column, max_row=sheet.max_row):
    for cell in row:
        if cell.data_type == 'f':  # Verifica se a célula contém uma fórmula
            print(f"Célula {cell.coordinate}: {cell.value}")
